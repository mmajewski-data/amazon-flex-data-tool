import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange

# Define all subtasks and columns in one place (GLOBAL) - Keywords removed
SUBTASKS = {
    "PD BP": [
        "asin", "gl_product_group_type.value", "product_type.value", "item_name.value",
        "brand.value", "item_description.value", "rtip_product_description.value",
        "bullet_point#1.value", "bullet_point#2.value", "bullet_point#3.value",
        "bullet_point#4.value", "bullet_point#5.value"
    ],
    "Attributes": [
        "asin", "item_name.value", "brand.value", "product_type.value",
        "gl_product_group_type.value", "color.value", "size.value", "model_name.value",
        "model_number.value", "item_type_name.value"
    ],
    "TCU RA": [
        "asin", "item_name.value", "product_type.value", "gl_product_group_type.value",
        "department.value", "part_number.value", "material#1.value", "flavor.value", "sub_brand.value",
        "hard_disk.description#1.value", "graphics_coprocessor.value", "operating_system#1.value", 
        "keyboard_layout.value", "voltage#1.value", "voltage#1.unit", "wattage.value", "wattage.unit",
        "cpu_model.manufacturer#1.value", "cpu_model.family#1.value", "cpu_model.model_number#1.value",
        "computer_memory.size#1.value", "computer_memory.size#1.unit", "memory_storage_capacity.value",
        "memory_storage_capacity.unit"
    ]
}

def gather_required_columns(subtasks):
    """
    Collect all unique columns needed by the subtasks.
    """
    all_columns = set()
    for columns in subtasks.values():
        all_columns.update(columns)
    return list(all_columns)

def create_master_file(input_file, master_file, subtasks):
    """
    Create a master file with the raw_data sheet populated ONLY with 
    the columns needed by the subtasks, thus improving performance.
    """
    # Gather all required columns from the subtasks
    required_columns = gather_required_columns(subtasks)

    # Step 1: Load ONLY the required columns into a DataFrame
    raw_data = pd.read_excel(input_file, sheet_name=0, usecols=required_columns)

    # Step 2: Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "raw_data"

    # Step 3: Write the DataFrame to the sheet
    for row in dataframe_to_rows(raw_data, index=False, header=True):
        ws.append(row)

    # Step 4: Save the workbook
    wb.save(master_file)
    print(f"Master file created with 'RAW DATA' sheet (only needed columns): {master_file}")

def create_subtask_from_raw(wb, raw_data, sheet_name, required_columns):
    """
    Create a new subtask sheet using data from the raw_data DataFrame.
    Args:
        wb (Workbook): The openpyxl workbook object.
        raw_data (DataFrame): The raw_data DataFrame loaded from the raw_data sheet.
        sheet_name (str): Name of the subtask sheet to create.
        required_columns (list): List of columns to include in the subtask sheet.
    """
    # Define additional blank columns specifically for "TCU RA" subtask
    additional_blank_columns = []
    if sheet_name == "TCU RA":
        additional_blank_columns = [
            "cpu_model.value", 
            "computer_memory.value", 
            "memory_storage_capacity.value"
        ]

    # Filter the DataFrame for required columns that exist in the raw data
    existing_columns = [col for col in required_columns if col in raw_data.columns]
    subtask_df = raw_data[existing_columns].copy()

    # Add additional blank columns to the DataFrame if applicable
    for blank_col in additional_blank_columns:
        if blank_col not in subtask_df.columns:
            subtask_df[blank_col] = ""  # Add a blank column with an empty string

    # Leave blank cells as empty strings
    subtask_df = subtask_df.where(pd.notnull(subtask_df), "")

    # Create a new sheet in the workbook
    ws = wb.create_sheet(title=sheet_name)

    # Write the DataFrame to the sheet
    for row in dataframe_to_rows(subtask_df, index=False, header=True):
        ws.append(row)

    print(f"Sheet '{sheet_name}' added to the master file.")

def create_title_cleanup_sheet(wb):
    """
    Add a Title Cleanup sheet to the workbook with an updated message.
    """
    ws = wb.create_sheet(title="Title Cleanup")
    ws.append([
        "Template for Papa Cleaner will be created together with the ready-to-upload SC files. "
        "It will be created for all ASINs regardless of their READY status."
    ])
    print("Sheet 'Title Cleanup' added to the master file.")


def set_column_width(sheet, default_width, col_a_width=8):
    """
    Set a default width for all columns starting from column B,
    and set a specific width for column A.

    Args:
        sheet (openpyxl worksheet): The worksheet to modify.
        default_width (float): The width to set for all columns starting from B.
        col_a_width (float): The width to set for column A.
    """
    # Set width for column A
    sheet.column_dimensions["A"].width = col_a_width

    # Set default width for columns starting from B
    for col in sheet.iter_cols(min_col=2):  # min_col=2 starts from column B
        col_letter = col[0].column_letter  # Get the column letter (e.g., 'B', 'C')
        sheet.column_dimensions[col_letter].width = default_width

def add_ready_column(wb, sheet_name):
    """
    Add a 'READY' column with a dropdown menu ('yes'/'no') to the specified sheet.
    Defaults to 'no' for all rows and strictly enforces the dropdown validation.
    """
    ws = wb[sheet_name]

    # Insert the 'READY' column at column A (first column)
    ws.insert_cols(1)
    ws["A1"] = "READY"  # Set the column header to 'READY'

    # Create a data validation for dropdown ('yes'/'no') and enforce validation
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=False, showErrorMessage=True)
    dv.error = "Invalid entry. Please select 'yes' or 'no' from the dropdown."
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Select 'yes' or 'no'."

    # Apply the dropdown to the range A2:A{num_rows}
    num_rows = ws.max_row
    for row in range(2, num_rows + 1):  # Skip header row
        ws[f"A{row}"] = "no"  # Set default value to 'no'

    # Add the data validation to the column range A2:A{num_rows}
    ws.add_data_validation(dv)
    dv.add(f"A2:A{num_rows}")  # Pass the range directly as a string

    print(f"'READY' column with dropdown added to '{sheet_name}'.")

def master_function(input_file, output_file, update_progress=None):
    """
    Create a master file with all subtasks and formatted sheets,
    using only the columns needed by SUBTASKS to improve performance.
    
    The optional 'update_progress' is a callback that accepts a float in [0.0, 1.0].
    """
    # Step 1: Create master file with ONLY needed columns
    create_master_file(input_file, output_file, SUBTASKS)
    if update_progress:
        update_progress(0.1)  # Example initial progress

    # Step 2: Load the master file once
    wb = load_workbook(output_file)

    # Step 3: Read 'raw_data' into a DataFrame
    raw_data_sheet = wb["raw_data"]
    raw_data = pd.DataFrame(raw_data_sheet.iter_rows(values_only=True))
    raw_data.columns = raw_data.iloc[0]  # Set the first row as columns
    raw_data = raw_data[1:]  # Remove the header row

    subtasks_count = len(SUBTASKS)
    total_steps = subtasks_count + 2
    current_step = 1  # after create_master_file

    # Step 4: Create all subtasks
    for subtask_name, columns in SUBTASKS.items():
        create_subtask_from_raw(
            wb, raw_data, subtask_name, columns
        )
        add_ready_column(wb, subtask_name)  # Add READY column with dropdown
        current_step += 1
        if update_progress:
            update_progress(current_step / total_steps)

    # Step 5: Add Title Cleanup sheet
    create_title_cleanup_sheet(wb)
    current_step += 1
    if update_progress:
        update_progress(current_step / total_steps)

    # Step 6: Set tab colors in batch
    for sheet_name in SUBTASKS.keys():
        wb[sheet_name].sheet_properties.tabColor = "006400"
    wb["Title Cleanup"].sheet_properties.tabColor = "006400"

    # Step 7: Adjust column widths
    for sheet_name in SUBTASKS.keys():
        set_column_width(wb[sheet_name], 25)

    # Step 8: Save the workbook once
    wb.save(output_file)
    current_step += 1
    if update_progress:
        update_progress(current_step / total_steps)

    print(f"Master file created at {output_file}")