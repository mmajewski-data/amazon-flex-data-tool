import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
import getpass
import os

# Configuration for subtasks with an additional "replace_nulls" flag - Keywords removed
SUBTASK_CONFIG = {
    "PD BP": {
        "column_map": {
            "asin": "sku",  # Changed from asin to sku
            "rtip_product_description.value": "rtip_product_description.value",
            "bullet_point#1.value": "bullet_point#1.value",
            "bullet_point#2.value": "bullet_point#2.value",
            "bullet_point#3.value": "bullet_point#3.value",
            "bullet_point#4.value": "bullet_point#4.value",
            "bullet_point#5.value": "bullet_point#5.value",
        },
        "default_values": {
            # Removed sc_vendor_name, added new fields
            "contributor_name": "EES-IDQ-DataAugmenter",
            "contributor_id": "835176407912",
            "merchant_suggested_asin.value": "",  # Will be filled with sku value
            "bullet_point#6.value": "",  # Changed from NULL to empty string
            "bullet_point#7.value": "",  # Changed from NULL to empty string
            "bullet_point#8.value": "",  # Changed from NULL to empty string
            "bullet_point#9.value": "",  # Changed from NULL to empty string
            "bullet_point#10.value": "",  # Changed from NULL to empty string
        },
        "replace_nulls": False,  # Changed from True to False
    },
    "Attributes": {
        "column_map": {
            "asin": "sku",  # Changed from asin to sku
            "color.value": "color.value",
            "size.value": "size.value",
            "model_number.value": "model_number.value",
            "model_name.value": "model_name.value",
            "item_type_name.value": "item_type_name.value"
        },
        "default_values": {
            # Removed sc_vendor_name, added new fields
            "contributor_name": "EES-IDQ-DataAugmenter",
            "contributor_id": "835176407912",
            "merchant_suggested_asin.value": "",  # Will be filled with sku value
        },
        "replace_nulls": False,  # Changed from True to False
    },
    "TCU RA": {
        "column_map": {
            "asin": "sku",  # Changed from asin to sku
            "department.value": "department.value",
            "part_number.value": "part_number.value",
            "material#1.value": "material#1.value",
            "flavor.value": "flavor.value",
            "sub_brand.value": "sub_brand.value",
            "hard_disk.description#1.value": "hard_disk.description#1.value",
            "graphics_coprocessor.value": "graphics_coprocessor.value",
            "operating_system#1.value": "operating_system#1.value",
            "keyboard_layout.value": "keyboard_layout.value"
        },
        "default_values": {
            # Removed sc_vendor_name, added new fields
            "contributor_name": "EES-IDQ-DataAugmenter",
            "contributor_id": "835176407912",
            "merchant_suggested_asin.value": "",  # Will be filled with sku value
        },
        "replace_nulls": False,
    },
}

# Utility function to generate file names with correct prefixes
def generate_file_name(subtask, qa_version=False):
    day = datetime.today().strftime("%m%d20%y")
    username = getpass.getuser()
    prefix_map = {
        "PD BP": "FLEX_ATTRPDB",
        "Attributes": "ATTR",
        "TCU RA": "FLEX_TCRA"
    }
    prefix = prefix_map.get(subtask, "FLEX_FILE")
    suffix = "_QA" if qa_version else ""
    return f"{prefix} {day}_{username}{suffix}.xlsx"

# Generate the ready-to-upload file for a given subtask
def create_ready_file(sheet_name, workbook_path, output_folder, column_map, default_values, replace_nulls=True):
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    filtered_df = df[df["READY"].str.lower() == "yes"].copy()

    if filtered_df.empty:
        print(f"No rows marked as READY in '{sheet_name}'. Skipping...")
        return

    renamed_df = filtered_df[list(column_map.keys())].rename(columns=column_map)

    for col, default_value in default_values.items():
        renamed_df[col] = default_value

    # Set merchant_suggested_asin.value to match the sku value
    if "sku" in renamed_df.columns and "merchant_suggested_asin.value" in renamed_df.columns:
        renamed_df["merchant_suggested_asin.value"] = renamed_df["sku"]

    # Reorder columns - put sku first, then contributor_name, contributor_id, merchant_suggested_asin.value
    columns = renamed_df.columns.tolist()
    for col in ["merchant_suggested_asin.value", "contributor_id", "contributor_name", "sku"]:
        if col in columns:
            columns.remove(col)
    
    # Insert in reverse order to maintain order
    columns.insert(0, "merchant_suggested_asin.value")
    columns.insert(0, "contributor_id")
    columns.insert(0, "contributor_name")
    columns.insert(0, "sku")
    
    renamed_df = renamed_df[columns]

    if replace_nulls:
        renamed_df = renamed_df.fillna("NULL")
    else:
        renamed_df = renamed_df.where(pd.notnull(renamed_df), "")

    renamed_df = renamed_df.astype(object)

    ready_file_name = generate_file_name(sheet_name)
    ready_file_path = os.path.join(output_folder, ready_file_name)
    renamed_df.to_excel(ready_file_path, index=False, header=True)

    print(f"Ready-to-upload file created: {ready_file_path}")
    add_version_to_file(ready_file_path)
    create_qa_file(renamed_df, sheet_name, output_folder)

def add_version_to_file(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # Insert ONE new row and add both version and marketplace information
    ws.insert_rows(1, 1)
    ws["A1"] = "version=1.0.0"
    ws["B1"] = "marketplace_id=712115121"

    # Define default font, border, and alignment
    default_font = Font(bold=False)
    default_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None)
    )
    default_alignment = Alignment(horizontal="left")

    # Apply styling to headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)  # Headers now at row 2 because we inserted only 1 row
        cell.font = default_font
        cell.border = default_border
        cell.alignment = default_alignment

    # Save the modified file
    wb.save(file_path)
    print(f"'version=1.0.0' added to A1, 'marketplace_id=712115121' added to B1, and headers styled: {file_path}")

def create_qa_file(dataframe, sheet_name, output_folder):
    qa_df = dataframe.copy()
    username = getpass.getuser()
    qa_df["login"] = username

    qa_file_name = generate_file_name(sheet_name, qa_version=True)
    qa_file_path = os.path.join(output_folder, qa_file_name)
    qa_df.to_excel(qa_file_path, index=False, header=True)

    print(f"QA file created: {qa_file_path}")

def create_title_cleanup_file(attributes_path, tcu_ra_path, output_folder):
    """
    Create a Title Cleanup file by merging data from Attributes and TCU RA spreadsheets.

    Args:
        attributes_path (str): Path to the Attributes sheet in the master workbook.
        tcu_ra_path (str): Path to the TCU RA sheet in the master workbook.
        output_folder (str): Folder to save the Title Cleanup file.
    """
    # Load data from Attributes and TCU RA sheets
    attributes_df = pd.read_excel(attributes_path, sheet_name="Attributes")
    tcu_ra_df = pd.read_excel(tcu_ra_path, sheet_name="TCU RA")

    # Rename columns as specified
    column_renames = {
        "material#1.value": "material.value",
        "voltage#1.value": "voltage.value",
        "hard_disk.description#1.value": "hard_disk.description.value",
        "operating_system#1.value": "operating_system.value",
    }
    attributes_df.rename(columns=column_renames, inplace=True)
    tcu_ra_df.rename(columns=column_renames, inplace=True)

    # List of required columns for the final file
    required_columns = [
        "asin", "brand.value", "color.value", "department.value",
        "gl_product_group_type.value", "item_type_name.value", "flavor.value",
        "material.value", "model_name.value", "model_number.value", "part_number.value",
        "product_type.value", "size.value", "wattage.value", "voltage.value",
        "sub_brand.value", "cpu_model.value", "computer_memory.value", 
        "memory_storage_capacity.value", "hard_disk.description.value", 
        "graphics_coprocessor.value", "operating_system.value", "keyboard_layout.value"
    ]

    # Perform a merge (like a VLOOKUP in Excel) on the "asin" column
    merged_df = pd.merge(
        attributes_df,
        tcu_ra_df,
        on="asin",
        how="outer",  # Use 'outer' to include all asins from both sheets
        suffixes=("", "_tcu_ra")  # Prevent column name conflicts
    )

    # Select and reorder the required columns
    final_df = merged_df[required_columns]

    # Explicitly convert columns to object dtype to prevent dtype conflicts
    final_df = final_df.astype(object)

    # Replace NaN with empty strings
    final_df = final_df.fillna("")

    # Generate the output file
    today = datetime.today().strftime("%m%d%Y")
    username = getpass.getuser()
    output_file = os.path.join(output_folder, f"papa_cleaner_template_{today}_{username}.xlsx")
    final_df.to_excel(output_file, index=False, header=True)

    print(f"Title Cleanup file created: {output_file}")


def generate_files(selected_tasks, workbook_path, output_folder):
    for task in selected_tasks:
        if task == "Title Cleanup":
            print(f"Generating file for task: {task}")
            create_title_cleanup_file(workbook_path, workbook_path, output_folder)
        else:
            config = SUBTASK_CONFIG.get(task)
            if config:
                print(f"Generating file for task: {task}")
                print(f"replace_nulls={config.get('replace_nulls', True)}")

                create_ready_file(
                    sheet_name=task,
                    workbook_path=workbook_path,
                    output_folder=output_folder,
                    column_map=config["column_map"],
                    default_values=config["default_values"],
                    replace_nulls=config.get("replace_nulls", True)
                )
            else:
                print(f"No configuration found for task: {task}")

    print("All files generated successfully!")